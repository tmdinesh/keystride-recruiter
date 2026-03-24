import pandas as pd

# Combined labels
labels_data = {
    "P046": "reject", "P024": "reject", "P044": "reject", "P047": "reject", "P025": "reject",
    "P020": "reject", "P016": "strong_fit", "P013": "reject", "P042": "reject", "P023": "reject",
    "P034": "reject", "P041": "reject", "P003": "reject", "P012": "reject", "P029": "strong_fit",
    "P045": "reject", "P026": "weak_fit", "P030": "medium_fit", "P043": "strong_fit", "P015": "reject",
    "P035": "reject", "P028": "reject", "P031": "reject", "P005": "reject", "P011": "reject",
    "P001": "medium_fit", "P002": "weak_fit", "P009": "reject", "P027": "reject", "P038": "reject",
    "P021": "medium_fit", "P006": "strong_fit", "P004": "reject", "P037": "reject", "P032": "medium_fit",
    "P033": "reject", "P017": "strong_fit", "P040": "medium_fit", "P022": "medium_fit", "P014": "reject",
    "P039": "reject", "P019": "reject", "P018": "reject", "P010": "reject", "P036": "reject",
    "P007": "strong_fit", "P008": "reject"
}

# Read the excel file
# We know it has 2 header rows to skip for data, but we want to preserve them for writing if possible.
# Actually, the best way to preserve formatting is to read with openpyxl or just read as is and write carefully.

import openpyxl

wb = openpyxl.load_workbook('ground_truth_labels.xlsx')
ws = wb.active

# Find the columns
header_row = 3 # Row 1 and 2 are instructions/titles, Row 3 is header
cols = {}
for cell in ws[header_row]:
    cols[cell.value] = cell.column

pair_id_col = cols.get('Pair ID')
fit_label_col = cols.get('Fit Label ▼')
labeled_by_col = cols.get('Labeled By')

if not all([pair_id_col, fit_label_col, labeled_by_col]):
    print(f"Error: Could not find columns. Pair ID: {pair_id_col}, Fit Label: {fit_label_col}, Labeled By: {labeled_by_col}")
    # Fallback to pandas if openpyxl fails to find headers correctly
else:
    # Iterate through rows starting from 4
    for row in range(4, ws.max_row + 1):
        pair_id = ws.cell(row=row, column=pair_id_col).value
        if pair_id in labels_data:
            ws.cell(row=row, column=fit_label_col).value = labels_data[pair_id]
            ws.cell(row=row, column=labeled_by_col).value = "Ollama CLI"

    wb.save('ground_truth_labels_filled.xlsx')
    print("Successfully updated ground_truth_labels_filled.xlsx")
