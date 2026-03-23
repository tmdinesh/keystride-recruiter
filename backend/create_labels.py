"""
Ground Truth Label Generator
Creates a formatted Excel sheet with auto-paired resume-JD combinations
"""
import os
import random
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

random.seed(42)

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
RESUMES_DIR = os.path.join(DATA_DIR, "resumes_anon")
JDS_DIR = os.path.join(DATA_DIR, "jds", "jds")
OUTPUT_PATH = os.path.join(DATA_DIR, "ground_truth_labels.xlsx")

ROLE_KEYWORDS = {
    "frontend":  ["frontend", "react", "vue", "angular", "javascript", "ui"],
    "backend":   ["backend", "node", "django", "flask", "java", "spring", "golang", "api"],
    "data":      ["data analyst", "data analysis", "sql", "tableau", "power bi", "analytics"],
    "ml":        ["machine learning", "ml", "deep learning", "nlp", "ai", "data scientist"],
    "qa":        ["qa", "quality", "test", "selenium", "automation", "sdet"],
    "devops":    ["devops", "cloud", "aws", "kubernetes", "docker", "sre", "platform"],
}

def detect_role(filename, text=""):
    combined = (filename + " " + text).lower()
    for role, keywords in ROLE_KEYWORDS.items():
        if any(kw in combined for kw in keywords):
            return role
    return "general"

def get_files(folder, exts):
    if not os.path.exists(folder):
        return []
    return [f for f in os.listdir(folder) if any(f.lower().endswith(e) for e in exts)]

def pair_resumes_jds():
    resume_files = get_files(RESUMES_DIR, [".txt"])
    jd_files     = get_files(JDS_DIR,      [".txt"])

    if not resume_files:
        resume_files = [f"resume_{i:03d}_anon.txt" for i in range(1, 56)]
    if not jd_files:
        jd_files = [f"jd_{i:03d}_{r}.txt"
                    for i, r in enumerate(["frontend","backend","data","ml","qa","devops"]*5, 1)]

    # Group by role
    resume_by_role = {r: [] for r in ROLE_KEYWORDS}
    resume_by_role["general"] = []
    for f in resume_files:
        resume_by_role[detect_role(f)].append(f)

    jd_by_role = {r: [] for r in ROLE_KEYWORDS}
    jd_by_role["general"] = []
    for f in jd_files:
        jd_by_role[detect_role(f)].append(f)

    pairs = []
    pair_id = 1

    # Role-matched pairs (strong/medium candidates)
    for role in ROLE_KEYWORDS:
        resumes = resume_by_role.get(role, []) + resume_by_role.get("general", [])
        jds     = jd_by_role.get(role, [])
        if not jds:
            continue
        for resume in resumes[:5]:
            jd = random.choice(jds)
            pairs.append({
                "pair_id":       f"P{pair_id:03d}",
                "resume_file":   resume,
                "jd_file":       jd,
                "role_category": role,
                "match_type":    "same_role",
                "fit_label":     "",
                "notes":         "",
            })
            pair_id += 1

    # Cross-role pairs (weak/reject candidates)
    all_resumes = resume_files.copy()
    random.shuffle(all_resumes)
    all_jds = jd_files.copy()
    random.shuffle(all_jds)

    for i in range(30):
        resume = all_resumes[i % len(all_resumes)]
        jd     = all_jds[(i + 3) % len(all_jds)]
        resume_role = detect_role(resume)
        jd_role     = detect_role(jd)
        if resume_role != jd_role:
            pairs.append({
                "pair_id":       f"P{pair_id:03d}",
                "resume_file":   resume,
                "jd_file":       jd,
                "role_category": f"{resume_role}→{jd_role}",
                "match_type":    "cross_role",
                "fit_label":     "",
                "notes":         "",
            })
            pair_id += 1

    random.shuffle(pairs)
    return pairs[:120]  # cap at 120 for the team to label


def build_excel(pairs, output_path):
    wb = Workbook()

    # ── Sheet 1: Labeling Sheet ───────────────────────────────────────────────
    ws = wb.active
    ws.title = "Label Pairs"

    # Colors
    HEADER_BG   = "1F3864"   # dark navy
    HEADER_FG   = "FFFFFF"
    SUBHEAD_BG  = "2E75B6"   # medium blue
    ALT_ROW     = "EBF3FB"   # light blue
    WHITE       = "FFFFFF"
    ACCENT      = "C6EFCE"   # light green for filled rows
    BORDER_COL  = "BDD7EE"

    thin = Side(style="thin", color=BORDER_COL)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "📋  Resume–JD Ground Truth Labeling Sheet"
    title_cell.font      = Font(name="Arial", bold=True, size=14, color=HEADER_FG)
    title_cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = (
        "Instructions: Read each Resume + JD pair and assign a Fit Label. "
        "Valid values: strong_fit | medium_fit | weak_fit | reject"
    )
    sub.font      = Font(name="Arial", size=10, italic=True, color=HEADER_FG)
    sub.fill      = PatternFill("solid", fgColor=SUBHEAD_BG)
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    # Headers
    headers = ["Pair ID", "Resume File", "JD File", "Role Category",
               "Match Type", "Fit Label ▼", "Labeled By", "Notes"]
    col_widths = [10, 32, 32, 18, 14, 18, 16, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = Font(name="Arial", bold=True, size=10, color=HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[3].height = 20

    # Data validation dropdown
    dv = DataValidation(
        type="list",
        formula1='"strong_fit,medium_fit,weak_fit,reject"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Label",
        error='Please choose: strong_fit, medium_fit, weak_fit, or reject'
    )
    ws.add_data_validation(dv)

    # Data rows
    for i, pair in enumerate(pairs):
        row = i + 4
        is_alt = (i % 2 == 0)
        row_fill = PatternFill("solid", fgColor=ALT_ROW if is_alt else WHITE)

        values = [
            pair["pair_id"],
            pair["resume_file"],
            pair["jd_file"],
            pair["role_category"],
            pair["match_type"],
            pair["fit_label"],
            "",   # labeled by
            pair["notes"],
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font      = Font(name="Arial", size=9)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col in (2, 3, 8)))
            if col == 6:  # Fit Label column
                cell.font = Font(name="Arial", size=9, bold=True)
                dv.add(cell)
        ws.row_dimensions[row].height = 18

    # Freeze panes
    ws.freeze_panes = "A4"

    # ── Sheet 2: Instructions ─────────────────────────────────────────────────
    wi = wb.create_sheet("Instructions")
    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 70

    wi.merge_cells("A1:B1")
    t = wi["A1"]
    t.value = "How to Label Resume–JD Pairs"
    t.font  = Font(name="Arial", bold=True, size=13, color=HEADER_FG)
    t.fill  = PatternFill("solid", fgColor=HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    wi.row_dimensions[1].height = 28

    guide = [
        ("Label",        "Criteria"),
        ("strong_fit",   "Candidate meets 80%+ of required skills, experience aligns, relevant projects"),
        ("medium_fit",   "Candidate meets 50–79% of required skills, some gaps but trainable"),
        ("weak_fit",     "Candidate meets <50% of required skills, significant experience mismatch"),
        ("reject",       "Completely unrelated role/skills, no overlap with JD requirements"),
        ("",             ""),
        ("Tips",         ""),
        ("✅ Focus on",  "Required skills match, years of experience, relevant project domain"),
        ("❌ Ignore",    "Candidate name, gender, location, university name, photo"),
        ("📌 Note",      "Each team member should label ~25 pairs. Aim for balanced distribution."),
        ("🎯 Target",    "25 strong_fit + 25 medium_fit + 25 weak_fit + 25 reject = 100 minimum"),
    ]

    label_colors = {
        "strong_fit": "C6EFCE", "medium_fit": "FFEB9C",
        "weak_fit": "FFCCCC",   "reject": "F4CCCC",
    }

    for r, (label, desc) in enumerate(guide, 2):
        c1 = wi.cell(row=r, column=1, value=label)
        c2 = wi.cell(row=r, column=2, value=desc)
        for c in (c1, c2):
            c.font      = Font(name="Arial", size=10, bold=(r == 2))
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border    = border
        if label in label_colors:
            fill = PatternFill("solid", fgColor=label_colors[label])
            c1.fill = fill
            c2.fill = fill
        wi.row_dimensions[r].height = 22

    # ── Sheet 3: Progress Tracker ─────────────────────────────────────────────
    wp = wb.create_sheet("Progress Tracker")
    wp.column_dimensions["A"].width = 22
    wp.column_dimensions["B"].width = 20
    wp.column_dimensions["C"].width = 20

    wp.merge_cells("A1:C1")
    pt = wp["A1"]
    pt.value = "Labeling Progress Tracker"
    pt.font  = Font(name="Arial", bold=True, size=13, color=HEADER_FG)
    pt.fill  = PatternFill("solid", fgColor=HEADER_BG)
    pt.alignment = Alignment(horizontal="center", vertical="center")
    wp.row_dimensions[1].height = 28

    progress_headers = ["Label", "Count (formula)", "Target"]
    for col, h in enumerate(progress_headers, 1):
        c = wp.cell(row=2, column=col, value=h)
        c.font      = Font(name="Arial", bold=True, size=10, color=HEADER_FG)
        c.fill      = PatternFill("solid", fgColor=SUBHEAD_BG)
        c.alignment = Alignment(horizontal="center")
        c.border    = border

    labels_targets = [
        ("strong_fit", 25), ("medium_fit", 25),
        ("weak_fit",   25), ("reject",     25), ("TOTAL", 100),
    ]
    last_data_row = 3 + len(pairs)

    for r, (lbl, target) in enumerate(labels_targets, 3):
        c1 = wp.cell(row=r, column=1, value=lbl)
        if lbl == "TOTAL":
            formula = f"=SUM(B3:B{r-1})"
        else:
            formula = f"=COUNTIF('Label Pairs'!F4:F{last_data_row},\"{lbl}\")"
        c2 = wp.cell(row=r, column=2, value=formula)
        c3 = wp.cell(row=r, column=3, value=target)

        for c in (c1, c2, c3):
            c.font      = Font(name="Arial", size=10, bold=(lbl == "TOTAL"))
            c.alignment = Alignment(horizontal="center")
            c.border    = border
        if lbl in label_colors:
            fill = PatternFill("solid", fgColor=label_colors[lbl])
            c1.fill = fill

    wb.save(output_path)
    print(f"✅ Excel file saved → {output_path}")
    print(f"   Total pairs generated: {len(pairs)}")


if __name__ == "__main__":
    os.makedirs(DATA_DIR, exist_ok=True)
    pairs = pair_resumes_jds()
    build_excel(pairs, OUTPUT_PATH)
